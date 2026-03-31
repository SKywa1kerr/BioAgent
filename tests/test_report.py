import pytest
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src-python"))

from bioagent.report import generate_excel_from_samples

def test_generate_excel_basic():
    samples = [
        {"id": "S1", "clone": "C1", "status": "ok", "reason": "",
         "identity": 1.0, "coverage": 0.95, "frameshift": False,
         "aa_changes_n": 0, "seq_length": 800, "sub": 0, "ins": 0, "dele": 0},
        {"id": "S2", "clone": "C2", "status": "wrong", "reason": "S334L",
         "identity": 0.98, "coverage": 0.90, "frameshift": False,
         "aa_changes_n": 1, "seq_length": 750, "sub": 3, "ins": 0, "dele": 0},
    ]
    data = generate_excel_from_samples(samples, source_path="/test/ab1")
    assert len(data) > 0
    from io import BytesIO
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "摘要" in wb.sheetnames
    assert "样本明细" in wb.sheetnames
    detail = wb["样本明细"]
    assert detail.cell(row=2, column=3).value == "ok"
    assert detail.cell(row=3, column=3).value == "wrong"

def test_empty_samples():
    data = generate_excel_from_samples([])
    assert len(data) > 0
