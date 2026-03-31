"""Excel report generation for BioAgent Desktop."""
from __future__ import annotations
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(bold=True, size=14)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_STATUS_FILL = {"ok": _FILL_GREEN, "wrong": _FILL_RED, "uncertain": _FILL_AMBER}


def generate_excel_from_samples(samples: list[dict], source_path: str = "") -> bytes:
    """Generate Excel report directly from sample dicts (no DB dependency).

    Args:
        samples: list of sample result dicts from analyze_folder()
        source_path: original AB1 directory path for the summary sheet

    Returns:
        xlsx file content as bytes
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb, samples, source_path)
    _write_detail(wb, samples)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(wb: openpyxl.Workbook, samples: list[dict], source_path: str) -> None:
    ws = wb.create_sheet("摘要")
    ws.merge_cells("A1:D1")
    ws["A1"] = "BioAgent 分析报告"
    ws["A1"].font = _FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")

    total = len(samples)
    ok = sum(1 for s in samples if s.get("status") == "ok")
    wrong = sum(1 for s in samples if s.get("status") == "wrong")
    uncertain = total - ok - wrong
    pass_rate = (ok / total * 100) if total > 0 else 0

    identities = [s["identity"] for s in samples if s.get("identity") is not None]
    coverages = [s["coverage"] for s in samples if s.get("coverage") is not None]

    rows = [
        ("来源路径", source_path),
        ("", ""),
        ("样本总数", total),
        ("合格数", ok),
        ("不合格数", wrong),
        ("不确定数", uncertain),
        ("合格率 (%)", round(pass_rate, 2)),
        ("平均 Identity", round(sum(identities) / len(identities), 4) if identities else "N/A"),
        ("平均 CDS Coverage", round(sum(coverages) / len(coverages), 4) if coverages else "N/A"),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40


def _write_detail(wb: openpyxl.Workbook, samples: list[dict]) -> None:
    ws = wb.create_sheet("样本明细")
    headers = ["样本ID", "克隆", "状态", "原因", "Identity", "Coverage",
               "移码突变", "AA变化数", "序列长度", "替换", "插入", "缺失"]
    keys = ["id", "clone", "status", "reason", "identity", "coverage",
            "frameshift", "aa_changes_n", "seq_length", "sub", "ins", "dele"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    for row_idx, s in enumerate(samples, 2):
        for col_idx, key in enumerate(keys, 1):
            value = s.get(key, "")
            if isinstance(value, bool):
                value = "是" if value else "否"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
        fill = _STATUS_FILL.get(s.get("status", ""))
        if fill:
            ws.cell(row=row_idx, column=3).fill = fill

    if samples:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(samples) + 1}"

    widths = [14, 10, 8, 25, 10, 10, 10, 10, 10, 8, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
