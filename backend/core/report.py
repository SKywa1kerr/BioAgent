"""Report data layer, Excel generation, and PDF report for BioAgent MAX."""
from __future__ import annotations

import base64
import json
import os
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.db.database import init_db, db_session
from backend.db.models import Analysis, Sample

ALL_MODULES = ["summary", "detail_table", "charts", "abnormal_details", "alignment", "thresholds"]
MAX_ALIGNMENT_SAMPLES = 20

# ── Conditional-formatting fills ──────────────────────────────────────

_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_TITLE = Font(bold=True, size=14)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_STATUS_FILL = {
    "ok": _FILL_GREEN,
    "wrong": _FILL_RED,
    "uncertain": _FILL_AMBER,
}


def _build_report_data(analysis_id: str) -> dict[str, Any]:
    """Read DB and return structured report data.

    Returns dict with keys: analysis, samples, thresholds, stats.
    Raises ValueError if analysis_id not found.
    """
    init_db()
    with db_session() as session:
        analysis = session.get(Analysis, analysis_id)
        if analysis is None:
            raise ValueError(f"Analysis '{analysis_id}' not found")

        # Extract analysis dict
        analysis_dict = {
            "id": analysis.id,
            "name": analysis.name,
            "source_type": analysis.source_type,
            "source_path": analysis.source_path,
            "status": analysis.status,
            "total": analysis.total,
            "ok_count": analysis.ok_count,
            "wrong_count": analysis.wrong_count,
            "uncertain_count": analysis.uncertain_count,
            "created_at": str(analysis.created_at) if analysis.created_at else "",
            "finished_at": str(analysis.finished_at) if analysis.finished_at else "",
        }

        # Parse thresholds
        thresholds = {}
        if analysis.config_snapshot:
            thresholds = json.loads(analysis.config_snapshot)

        # Fetch samples
        samples_orm = session.query(Sample).filter(
            Sample.analysis_id == analysis_id
        ).order_by(Sample.sid).all()

        samples = []
        for s in samples_orm:
            samples.append({
                "id": s.id,
                "sid": s.sid,
                "clone": s.clone,
                "status": s.status,
                "reason": s.reason,
                "rule_id": s.rule_id,
                "identity": s.identity,
                "cds_coverage": s.cds_coverage,
                "frameshift": s.frameshift,
                "aa_changes_n": s.aa_changes_n,
                "raw_aa_changes_n": s.raw_aa_changes_n,
                "orientation": s.orientation,
                "seq_length": s.seq_length,
                "ref_length": s.ref_length,
                "avg_quality": s.avg_quality,
                "sub_count": s.sub_count,
                "ins_count": s.ins_count,
                "del_count": s.del_count,
            })

        # Compute stats
        total = analysis.total or 0
        ok = analysis.ok_count or 0
        pass_rate = (ok / total * 100) if total > 0 else 0.0

        identities = [s["identity"] for s in samples if s["identity"] is not None]
        coverages = [s["cds_coverage"] for s in samples if s["cds_coverage"] is not None]

        stats = {
            "pass_rate": round(pass_rate, 2),
            "avg_identity": round(sum(identities) / len(identities), 4) if identities else 0.0,
            "avg_coverage": round(sum(coverages) / len(coverages), 4) if coverages else 0.0,
        }

    return {
        "analysis": analysis_dict,
        "samples": samples,
        "thresholds": thresholds,
        "stats": stats,
    }


def generate_excel(analysis_id: str, modules: list[str] | None = None) -> bytes:
    """Generate a multi-sheet Excel report as bytes.

    Args:
        analysis_id: The analysis to report on.
        modules: Which sheets to include. None means all Excel-relevant modules.
                 Recognized: "summary" → 摘要, "detail_table" → 样本明细, "thresholds" → 阈值配置.

    Returns:
        xlsx file content as bytes.
    """
    if modules is None:
        modules = ["summary", "detail_table", "thresholds"]

    data = _build_report_data(analysis_id)

    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    if "summary" in modules:
        _write_summary_sheet(wb, data)

    if "detail_table" in modules:
        _write_detail_sheet(wb, data)

    if "thresholds" in modules:
        _write_thresholds_sheet(wb, data)

    # If no sheets were created (empty modules list), add a placeholder
    if len(wb.sheetnames) == 0:
        wb.create_sheet("空")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet writers ─────────────────────────────────────────────────────

def _write_summary_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    ws = wb.create_sheet("摘要")
    a = data["analysis"]
    stats = data["stats"]

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "分析报告摘要"
    ws["A1"].font = _FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("分析名称", a["name"]),
        ("来源类型", a["source_type"]),
        ("来源路径", a["source_path"]),
        ("状态", a["status"]),
        ("创建时间", a["created_at"]),
        ("完成时间", a["finished_at"]),
        ("", ""),
        ("样本总数", a["total"]),
        ("合格数", a["ok_count"]),
        ("不合格数", a["wrong_count"]),
        ("不确定数", a["uncertain_count"]),
        ("合格率 (%)", stats["pass_rate"]),
        ("平均 Identity", stats["avg_identity"]),
        ("平均 CDS Coverage", stats["avg_coverage"]),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40


def _write_detail_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    ws = wb.create_sheet("样本明细")
    samples = data["samples"]

    headers = [
        "样本ID", "克隆", "状态", "原因", "规则",
        "Identity", "CDS Coverage", "移码突变",
        "AA变化数", "方向", "序列长度", "参考长度",
        "平均质量", "替换数", "插入数", "缺失数",
    ]
    sample_keys = [
        "sid", "clone", "status", "reason", "rule_id",
        "identity", "cds_coverage", "frameshift",
        "aa_changes_n", "orientation", "seq_length", "ref_length",
        "avg_quality", "sub_count", "ins_count", "del_count",
    ]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # Write data rows
    for row_idx, sample in enumerate(samples, start=2):
        for col_idx, key in enumerate(sample_keys, start=1):
            value = sample.get(key, "")
            if isinstance(value, bool):
                value = "是" if value else "否"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER

        # Conditional formatting for status column (column 3)
        status = sample.get("status", "")
        fill = _STATUS_FILL.get(status)
        if fill:
            ws.cell(row=row_idx, column=3).fill = fill

    # Auto-filter
    if samples:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(samples) + 1}"

    # Column widths
    col_widths = [14, 10, 8, 25, 6, 10, 12, 10, 10, 10, 10, 10, 10, 8, 8, 8]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_thresholds_sheet(wb: openpyxl.Workbook, data: dict) -> None:
    ws = wb.create_sheet("阈值配置")
    thresholds = data["thresholds"]

    # Header row
    for col_idx, header in enumerate(["参数名称", "参数值"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    for row_idx, (key, value) in enumerate(sorted(thresholds.items()), start=2):
        ws.cell(row=row_idx, column=1, value=key).border = _THIN_BORDER
        ws.cell(row=row_idx, column=2, value=value).border = _THIN_BORDER

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


# ── Chart rendering ──────────────────────────────────────────────────

STATUS_COLORS = {"ok": "#16a34a", "wrong": "#dc2626", "uncertain": "#d97706"}


def _render_charts(samples: list[dict]) -> dict[str, bytes]:
    """Generate Plotly charts as PNG bytes via kaleido.

    Returns dict mapping chart name to PNG bytes:
      - identity_hist: Identity distribution histogram, colored by status
      - coverage_hist: CDS coverage distribution histogram, colored by status
      - status_pie: Status pie chart (donut style)
    """
    import plotly.graph_objects as go

    chart_width = 500
    chart_height = 350
    layout_defaults = dict(
        width=chart_width,
        height=chart_height,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Arial"),
        margin=dict(l=50, r=30, t=40, b=40),
    )

    charts: dict[str, bytes] = {}

    # Group samples by status
    status_groups: dict[str, list[dict]] = {}
    for s in samples:
        st = s.get("status", "uncertain")
        status_groups.setdefault(st, []).append(s)

    # 1. Identity histogram
    fig_id = go.Figure()
    for status in ["ok", "wrong", "uncertain"]:
        group = status_groups.get(status, [])
        vals = [s["identity"] for s in group if s.get("identity") is not None]
        if vals:
            fig_id.add_trace(go.Histogram(
                x=vals, name=status,
                marker_color=STATUS_COLORS.get(status, "#888"),
                opacity=0.8,
            ))
    fig_id.update_layout(
        title="Identity Distribution",
        xaxis_title="Identity",
        yaxis_title="Count",
        barmode="overlay",
        **layout_defaults,
    )
    charts["identity_hist"] = fig_id.to_image(format="png")

    # 2. Coverage histogram
    fig_cov = go.Figure()
    for status in ["ok", "wrong", "uncertain"]:
        group = status_groups.get(status, [])
        vals = [s["cds_coverage"] for s in group if s.get("cds_coverage") is not None]
        if vals:
            fig_cov.add_trace(go.Histogram(
                x=vals, name=status,
                marker_color=STATUS_COLORS.get(status, "#888"),
                opacity=0.8,
            ))
    fig_cov.update_layout(
        title="CDS Coverage Distribution",
        xaxis_title="CDS Coverage",
        yaxis_title="Count",
        barmode="overlay",
        **layout_defaults,
    )
    charts["coverage_hist"] = fig_cov.to_image(format="png")

    # 3. Status pie (donut)
    status_counts: dict[str, int] = {}
    for s in samples:
        st = s.get("status", "uncertain")
        status_counts[st] = status_counts.get(st, 0) + 1

    labels = list(status_counts.keys())
    values = list(status_counts.values())
    colors = [STATUS_COLORS.get(l, "#888") for l in labels]

    fig_pie = go.Figure(data=[go.Pie(
        labels=labels, values=values,
        hole=0.45,
        marker=dict(colors=colors),
    )])
    fig_pie.update_layout(
        title="Status Distribution",
        **layout_defaults,
    )
    charts["status_pie"] = fig_pie.to_image(format="png")

    return charts


# ── Alignment HTML rendering ─────────────────────────────────────────

def _render_alignment_html(ref_gapped: str, qry_gapped: str, width: int = 80) -> str:
    """Render gapped alignment as HTML with mismatch highlighting.

    Returns HTML string with <pre> blocks, no Streamlit dependency.
    Mismatches are highlighted with colored <span> tags.
    """
    if not ref_gapped or not qry_gapped:
        return "<pre>No alignment data</pre>"

    lines: list[str] = []
    length = min(len(ref_gapped), len(qry_gapped))

    for start in range(0, length, width):
        end = min(start + width, length)
        ref_chunk = ref_gapped[start:end]
        qry_chunk = qry_gapped[start:end]

        # Build match line and highlighted sequences
        ref_html_parts: list[str] = []
        qry_html_parts: list[str] = []
        match_parts: list[str] = []

        for r, q in zip(ref_chunk, qry_chunk):
            if r == q:
                ref_html_parts.append(r)
                qry_html_parts.append(q)
                match_parts.append("|")
            elif r == "-" or q == "-":
                ref_html_parts.append(f'<span style="color:#d97706;font-weight:bold">{r}</span>')
                qry_html_parts.append(f'<span style="color:#d97706;font-weight:bold">{q}</span>')
                match_parts.append(" ")
            else:
                ref_html_parts.append(f'<span style="color:#dc2626;font-weight:bold">{r}</span>')
                qry_html_parts.append(f'<span style="color:#dc2626;font-weight:bold">{q}</span>')
                match_parts.append("X")

        pos_label = f"{start + 1:>6}"
        lines.append(f"Ref {pos_label}  {''.join(ref_html_parts)}")
        lines.append(f"         {''.join(match_parts)}")
        lines.append(f"Qry {pos_label}  {''.join(qry_html_parts)}")
        lines.append("")

    return "<pre style=\"font-family:monospace;font-size:11px;line-height:1.4;white-space:pre-wrap\">" + "\n".join(lines) + "</pre>"


# ── HTML rendering ───────────────────────────────────────────────────

def _render_html(report_data: dict, modules: list[str]) -> str:
    """Load Jinja2 template and render HTML for PDF generation."""
    from datetime import datetime
    from jinja2 import Environment, FileSystemLoader

    template_dir = Path(__file__).resolve().parent.parent / "templates"
    env = Environment(loader=FileSystemLoader(str(template_dir)), autoescape=True)
    template = env.get_template("report.html")

    analysis = report_data["analysis"]
    samples = report_data["samples"]
    stats = report_data["stats"]
    thresholds = report_data["thresholds"]

    # Render charts and convert to base64
    charts_b64: dict[str, str] = {}
    if "charts" in modules:
        charts_raw = _render_charts(samples)
        for name, png_bytes in charts_raw.items():
            charts_b64[name] = base64.b64encode(png_bytes).decode("ascii")

    # Filter abnormal samples
    abnormal_samples: list[dict] = []
    for s in samples:
        if s.get("status") in ("wrong", "uncertain"):
            s_copy = dict(s)
            # Parse aa_changes if stored as JSON string
            aa_raw = s.get("aa_changes", "[]")
            if isinstance(aa_raw, str):
                try:
                    s_copy["aa_changes_parsed"] = json.loads(aa_raw)
                except (json.JSONDecodeError, TypeError):
                    s_copy["aa_changes_parsed"] = []
            elif isinstance(aa_raw, list):
                s_copy["aa_changes_parsed"] = aa_raw
            else:
                s_copy["aa_changes_parsed"] = []
            abnormal_samples.append(s_copy)

    # Render alignment HTML for abnormal samples
    alignment_samples: list[dict] = []
    alignment_truncated = False
    if "alignment" in modules:
        candidates = [s for s in samples if s.get("status") in ("wrong", "uncertain")]
        if len(candidates) > MAX_ALIGNMENT_SAMPLES:
            alignment_truncated = True
            candidates = candidates[:MAX_ALIGNMENT_SAMPLES]
        for s in candidates:
            ref_g = s.get("ref_gapped") or ""
            qry_g = s.get("qry_gapped") or ""
            alignment_html = _render_alignment_html(ref_g, qry_g)
            alignment_samples.append({
                "sid": s.get("sid", ""),
                "status": s.get("status", ""),
                "alignment_html": alignment_html,
            })

    context = {
        "analysis": analysis,
        "samples": samples,
        "stats": stats,
        "thresholds": thresholds,
        "charts": charts_b64,
        "abnormal_samples": abnormal_samples,
        "alignment_samples": alignment_samples,
        "alignment_truncated": alignment_truncated,
        "total_abnormal": len([s for s in samples if s.get("status") in ("wrong", "uncertain")]),
        "max_alignment": MAX_ALIGNMENT_SAMPLES,
        "modules": modules,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    return template.render(**context)


# ── PDF generation ───────────────────────────────────────────────────

def generate_pdf(analysis_id: str, modules: list[str] | None = None) -> bytes:
    """Generate a PDF report as bytes.

    Args:
        analysis_id: The analysis to report on.
        modules: Which sections to include. None means all modules.

    Returns:
        PDF file content as bytes.
    """
    try:
        from xhtml2pdf import pisa
    except ImportError:
        raise ImportError(
            "xhtml2pdf is required for PDF generation. "
            "Install it with: pip install xhtml2pdf"
        )

    if modules is None:
        modules = list(ALL_MODULES)

    data = _build_report_data(analysis_id)
    html_string = _render_html(data, modules)

    result_buf = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=result_buf)

    if pisa_status.err:
        raise RuntimeError(f"PDF generation failed with {pisa_status.err} error(s)")

    return result_buf.getvalue()
